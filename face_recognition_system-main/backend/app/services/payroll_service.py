# backend/app/services/payroll_service.py

from datetime import datetime, timedelta
import pytz
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, numbers
import os
from app.db import db
from app.models.payroll import PayrollConfig, PayrollRecord
from app.models.employee import Employee
from app.services.report_service import (
    process_employee_attendance_enhanced,
    get_working_days_between
)
from app.utils.helpers import get_vn_datetime, get_export_path


def _get_month_range(month, year):
    """Tạo datetime range cho tháng"""
    start = datetime(year, month, 1)
    if month == 12:
        end = datetime(year + 1, 1, 1) - timedelta(seconds=1)
    else:
        end = datetime(year, month + 1, 1) - timedelta(seconds=1)
    return start, end


def calculate_employee_salary(employee_id, month, year):
    """Tính lương cho 1 nhân viên trong tháng"""
    employee = Employee.query.filter_by(employee_id=employee_id.upper(), status=True).first()
    if not employee:
        return None, "Employee not found"
    
    config = PayrollConfig.get_current()
    start_date, end_date = _get_month_range(month, year)
    
    # Lấy dữ liệu chấm công từ report_service
    report = process_employee_attendance_enhanced(employee_id, start_date, end_date)
    
    # Tính toán
    working_days = report["working_days_in_period"]
    actual_days = report["total_attendance_days"]
    late_count = report["late_days"]
    absent_days = report["absent_days"]
    ot_hours = report["total_overtime_hours"]
    
    # Lương
    daily_rate = config.base_salary / config.working_days_per_month
    hourly_rate = daily_rate / 8  # 8 giờ/ngày
    
    work_salary = daily_rate * actual_days  # Lương theo ngày công
    ot_pay = hourly_rate * config.ot_multiplier * ot_hours  # Tiền OT
    late_deductions = config.late_deduction * late_count  # Trừ đi trễ
    absent_deductions = daily_rate * config.absent_deduction_rate * absent_days  # Trừ vắng
    
    gross_salary = work_salary + ot_pay
    total_deductions = late_deductions + absent_deductions
    net_salary = max(0, gross_salary - total_deductions)
    
    return {
        "employee_id": employee.employee_id,
        "full_name": employee.full_name,
        "department": employee.department,
        "position": employee.position,
        "month": month,
        "year": year,
        "working_days_standard": working_days,
        "actual_work_days": actual_days,
        "late_count": late_count,
        "absent_days": absent_days,
        "ot_hours": round(ot_hours, 2),
        "base_salary": config.base_salary,
        "daily_rate": round(daily_rate, 0),
        "work_salary": round(work_salary, 0),
        "ot_pay": round(ot_pay, 0),
        "late_deductions": round(late_deductions, 0),
        "absent_deductions": round(absent_deductions, 0),
        "gross_salary": round(gross_salary, 0),
        "total_deductions": round(total_deductions, 0),
        "net_salary": round(net_salary, 0)
    }, None


def generate_payroll_batch(month, year):
    """Tạo bảng lương cho tất cả NV active"""
    employees = Employee.query.filter_by(status=True).all()
    if not employees:
        return None, "No active employees found"
    
    config = PayrollConfig.get_current()
    results = []
    errors = []
    
    for emp in employees:
        try:
            salary_data, error = calculate_employee_salary(emp.employee_id, month, year)
            if error:
                errors.append({"employee_id": emp.employee_id, "error": error})
                continue
            
            # Kiểm tra record đã tồn tại chưa
            existing = PayrollRecord.query.filter_by(
                employee_id=emp.employee_id,
                month=month,
                year=year
            ).first()
            
            if existing:
                # Cập nhật
                existing.working_days_standard = salary_data["working_days_standard"]
                existing.actual_work_days = salary_data["actual_work_days"]
                existing.late_count = salary_data["late_count"]
                existing.absent_days = salary_data["absent_days"]
                existing.ot_hours = salary_data["ot_hours"]
                existing.base_salary = salary_data["base_salary"]
                existing.daily_rate = salary_data["daily_rate"]
                existing.work_salary = salary_data["work_salary"]
                existing.ot_pay = salary_data["ot_pay"]
                existing.late_deductions = salary_data["late_deductions"]
                existing.absent_deductions = salary_data["absent_deductions"]
                existing.gross_salary = salary_data["gross_salary"]
                existing.total_deductions = salary_data["total_deductions"]
                existing.net_salary = salary_data["net_salary"]
                existing.status = 'draft'
                record = existing
            else:
                # Tạo mới
                record = PayrollRecord(
                    employee_id=emp.employee_id,
                    month=month,
                    year=year,
                    working_days_standard=salary_data["working_days_standard"],
                    actual_work_days=salary_data["actual_work_days"],
                    late_count=salary_data["late_count"],
                    absent_days=salary_data["absent_days"],
                    ot_hours=salary_data["ot_hours"],
                    base_salary=salary_data["base_salary"],
                    daily_rate=salary_data["daily_rate"],
                    work_salary=salary_data["work_salary"],
                    ot_pay=salary_data["ot_pay"],
                    late_deductions=salary_data["late_deductions"],
                    absent_deductions=salary_data["absent_deductions"],
                    gross_salary=salary_data["gross_salary"],
                    total_deductions=salary_data["total_deductions"],
                    net_salary=salary_data["net_salary"],
                    status='draft'
                )
                db.session.add(record)
            
            results.append({**salary_data, "status": "draft"})
            
        except Exception as e:
            errors.append({"employee_id": emp.employee_id, "error": str(e)})
    
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return None, f"Database error: {str(e)}"
    
    return {
        "month": month,
        "year": year,
        "total_employees": len(employees),
        "processed": len(results),
        "errors": len(errors),
        "records": results,
        "error_details": errors
    }, None


def get_payroll_list(month, year):
    """Lấy danh sách lương tháng"""
    records = PayrollRecord.query.filter_by(month=month, year=year).all()
    
    result = []
    for record in records:
        emp = Employee.query.filter_by(employee_id=record.employee_id).first()
        data = record.to_dict()
        if emp:
            data["full_name"] = emp.full_name
            data["department"] = emp.department
            data["position"] = emp.position
        result.append(data)
    
    # Tổng kết
    total_gross = sum(r["gross_salary"] for r in result)
    total_deductions = sum(r["total_deductions"] for r in result)
    total_net = sum(r["net_salary"] for r in result)
    
    return {
        "month": month,
        "year": year,
        "records": result,
        "total_employees": len(result),
        "summary": {
            "total_gross": round(total_gross, 0),
            "total_deductions": round(total_deductions, 0),
            "total_net": round(total_net, 0)
        }
    }


def export_payroll_excel(month, year):
    """Xuất bảng lương ra Excel"""
    payroll_data = get_payroll_list(month, year)
    
    if not payroll_data["records"]:
        return None, "No payroll records found. Generate payroll first."
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Payroll {month:02d}/{year}"
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    money_format = '#,##0'
    
    # Title
    ws.append([f"BẢNG LƯƠNG THÁNG {month:02d}/{year}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_align
    ws.append([])
    
    # Headers
    headers = [
        "STT", "Mã NV", "Họ tên", "Phòng ban", "Ngày công", "Ngày nghỉ",
        "Số lần trễ", "Giờ OT", "Lương cơ bản", "Lương công",
        "Tiền OT", "Trừ đi trễ", "Trừ vắng", "Tổng trừ", "Thực nhận"
    ]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # Data rows
    for idx, record in enumerate(payroll_data["records"], 1):
        row = [
            idx,
            record["employee_id"],
            record.get("full_name", ""),
            record.get("department", ""),
            record["actual_work_days"],
            record["absent_days"],
            record["late_count"],
            record["ot_hours"],
            record["base_salary"],
            record["work_salary"],
            record["ot_pay"],
            record["late_deductions"],
            record["absent_deductions"],
            record["total_deductions"],
            record["net_salary"]
        ]
        ws.append(row)
        
        # Format money columns
        for col_idx in range(9, 16):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.number_format = money_format
    
    # Summary row
    ws.append([])
    summary_row = ws.max_row + 1
    ws.append(["", "", "", "TỔNG CỘNG", "", "", "", "",
               "", "", "", "", "",
               payroll_data["summary"]["total_deductions"],
               payroll_data["summary"]["total_net"]])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for col_idx in [14, 15]:
        ws.cell(row=ws.max_row, column=col_idx).number_format = money_format
    
    # Column widths
    widths = [5, 10, 25, 18, 10, 10, 10, 8, 15, 15, 12, 12, 12, 12, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    
    # Save
    export_path = get_export_path()
    os.makedirs(export_path, exist_ok=True)
    timestamp = get_vn_datetime().strftime("%Y%m%d_%H%M%S")
    filename = f"payroll_{month:02d}_{year}_{timestamp}.xlsx"
    filepath = os.path.join(export_path, filename)
    wb.save(filepath)
    
    return {
        "message": "Payroll exported successfully",
        "download_url": f"/api/reports/download/{filename}",
        "filename": filename,
        "file_path": filepath,
        "file_size": os.path.getsize(filepath)
    }, None
